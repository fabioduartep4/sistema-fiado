"""Testes de integração do serviço de relatórios e painel de início.

Gravam de verdade no banco configurado em ``.env``. Só rodam com
``RODAR_TESTES_INTEGRACAO=1`` (ver ``tests/conftest.py``).
"""

from __future__ import annotations

import csv
import uuid
from datetime import timedelta
from decimal import Decimal

import openpyxl
import pytest

from app.services import cliente_service, compra_service, relatorio_service, usuario_service
from app.services.usuario_service import PermissaoNegadaError
from app.utils.date_utils import obter_data_padrao

pytestmark = pytest.mark.integration


def _criar_funcionario_teste(usuario_admin_teste):
    from app.models.usuario import PerfilUsuario
    from app.services import auth_service

    login = f"teste_func_relatorio_{uuid.uuid4().hex[:10]}"
    funcionario = usuario_service.criar_usuario(
        usuario_admin_teste, "Funcionário de Teste Relatório", login, "senha-func-123",
        PerfilUsuario.FUNCIONARIO,
    )
    usuario_funcionario = auth_service.autenticar(login, "senha-func-123")
    return funcionario, usuario_funcionario


def test_listar_historico_funcionario_e_rejeitado(usuario_admin_teste) -> None:
    funcionario, usuario_funcionario = _criar_funcionario_teste(usuario_admin_teste)

    with pytest.raises(PermissaoNegadaError):
        relatorio_service.listar_historico(usuario_funcionario)

    usuario_service.definir_ativo(usuario_admin_teste, funcionario.id, False)


def test_listar_log_erros_funcionario_e_rejeitado(usuario_admin_teste) -> None:
    funcionario, usuario_funcionario = _criar_funcionario_teste(usuario_admin_teste)

    with pytest.raises(PermissaoNegadaError):
        relatorio_service.listar_log_erros(usuario_funcionario)

    usuario_service.definir_ativo(usuario_admin_teste, funcionario.id, False)


def test_listar_saldos_em_aberto_funcionario_e_rejeitado(usuario_admin_teste) -> None:
    funcionario, usuario_funcionario = _criar_funcionario_teste(usuario_admin_teste)

    with pytest.raises(PermissaoNegadaError):
        relatorio_service.listar_saldos_em_aberto(usuario_funcionario)

    usuario_service.definir_ativo(usuario_admin_teste, funcionario.id, False)


def test_listar_saldos_em_atraso_funcionario_e_rejeitado(usuario_admin_teste) -> None:
    funcionario, usuario_funcionario = _criar_funcionario_teste(usuario_admin_teste)

    with pytest.raises(PermissaoNegadaError):
        relatorio_service.listar_saldos_em_atraso(usuario_funcionario)

    usuario_service.definir_ativo(usuario_admin_teste, funcionario.id, False)


def test_listar_saldos_em_atraso_inclui_compra_antiga_e_exclui_recente(usuario_admin_teste) -> None:
    hoje = obter_data_padrao()

    cliente_atrasado = cliente_service.cadastrar_cliente(
        usuario_admin_teste,
        "Teste Automatizado Relatorio Atraso",
        [],
        ["(35) 99999-1234"],
        [],
    )
    compra_service.registrar_compra(
        usuario_admin_teste, cliente_atrasado.id, Decimal("90.00"), hoje - timedelta(days=40), None
    )

    cliente_recente = cliente_service.cadastrar_cliente(
        usuario_admin_teste, "Teste Automatizado Relatorio Recente", [], [], []
    )
    compra_service.registrar_compra(
        usuario_admin_teste, cliente_recente.id, Decimal("20.00"), hoje, None
    )

    atrasados = relatorio_service.listar_saldos_em_atraso(usuario_admin_teste, dias_atraso=30)
    nomes_atrasados = [s.nome_principal for s in atrasados]

    assert cliente_atrasado.nome_principal in nomes_atrasados
    assert cliente_recente.nome_principal not in nomes_atrasados

    saldo_atrasado = next(s for s in atrasados if s.nome_principal == cliente_atrasado.nome_principal)
    assert saldo_atrasado.total_em_atraso == Decimal("90.00")
    assert saldo_atrasado.dias_desde_a_compra_mais_antiga >= 40
    assert saldo_atrasado.telefone is not None
    assert "99999" in saldo_atrasado.telefone

    cliente_service.excluir_cliente(usuario_admin_teste, cliente_atrasado.id)
    cliente_service.excluir_cliente(usuario_admin_teste, cliente_recente.id)


def test_listar_saldos_em_aberto_reflete_compra_em_aberto(usuario_admin_teste) -> None:
    cliente = cliente_service.cadastrar_cliente(
        usuario_admin_teste, "Teste Automatizado Relatorio Saldo", [], [], []
    )
    hoje = obter_data_padrao()
    compra_service.registrar_compra(usuario_admin_teste, cliente.id, Decimal("48.00"), hoje, None)

    saldos = relatorio_service.listar_saldos_em_aberto(usuario_admin_teste)
    saldo_cliente = next((s for s in saldos if s.nome_principal == cliente.nome_principal), None)

    assert saldo_cliente is not None
    assert saldo_cliente.total_em_aberto == Decimal("48.00")

    cliente_service.excluir_cliente(usuario_admin_teste, cliente.id)


def test_exportar_saldos_em_aberto_csv_gera_arquivo_correto(usuario_admin_teste, tmp_path) -> None:
    cliente = cliente_service.cadastrar_cliente(
        usuario_admin_teste, "Teste Automatizado Relatorio CSV", [], [], []
    )
    hoje = obter_data_padrao()
    compra_service.registrar_compra(usuario_admin_teste, cliente.id, Decimal("12.34"), hoje, None)

    caminho_csv = tmp_path / "saldo_em_aberto.csv"
    relatorio_service.exportar_saldos_em_aberto_csv(usuario_admin_teste, str(caminho_csv))

    assert caminho_csv.exists()
    with open(caminho_csv, encoding="utf-8-sig", newline="") as arquivo:
        linhas = list(csv.reader(arquivo, delimiter=";"))

    assert linhas[0] == ["Código", "Cliente", "Total em Aberto (R$)"]
    linha_cliente = next((l for l in linhas[1:] if l[1] == cliente.nome_principal), None)
    assert linha_cliente is not None
    assert linha_cliente[2] == "12.34"

    cliente_service.excluir_cliente(usuario_admin_teste, cliente.id)


def test_exportar_saldos_em_aberto_xlsx_gera_planilha_formatada(usuario_admin_teste, tmp_path) -> None:
    cliente = cliente_service.cadastrar_cliente(
        usuario_admin_teste, "Teste Automatizado Relatorio XLSX", [], [], []
    )
    hoje = obter_data_padrao()
    compra_service.registrar_compra(usuario_admin_teste, cliente.id, Decimal("56.70"), hoje, None)

    caminho_xlsx = tmp_path / "saldo_em_aberto.xlsx"
    relatorio_service.exportar_saldos_em_aberto_xlsx(usuario_admin_teste, str(caminho_xlsx))

    assert caminho_xlsx.exists()
    pasta_trabalho = openpyxl.load_workbook(caminho_xlsx)
    planilha = pasta_trabalho.active

    cabecalho = [celula.value for celula in planilha[1]]
    assert cabecalho == ["Código", "Cliente", "Total em Aberto (R$)"]
    assert planilha["A1"].font.bold is True

    linha_cliente = next(
        (linha for linha in planilha.iter_rows(min_row=2) if linha[1].value == cliente.nome_principal),
        None,
    )
    assert linha_cliente is not None
    assert linha_cliente[2].value == pytest.approx(56.70)

    ultima_linha = list(planilha.iter_rows(min_row=2))[-1]
    assert ultima_linha[0].value == "Total"

    cliente_service.excluir_cliente(usuario_admin_teste, cliente.id)


def test_obter_painel_inicio_funcionario_e_rejeitado(usuario_admin_teste) -> None:
    funcionario, usuario_funcionario = _criar_funcionario_teste(usuario_admin_teste)

    with pytest.raises(PermissaoNegadaError):
        relatorio_service.obter_painel_inicio(usuario_funcionario)

    usuario_service.definir_ativo(usuario_admin_teste, funcionario.id, False)


def test_obter_painel_inicio_total_em_aberto_reflete_nova_compra(usuario_admin_teste) -> None:
    cliente = cliente_service.cadastrar_cliente(
        usuario_admin_teste, "Teste Automatizado Relatorio Painel", [], [], []
    )
    hoje = obter_data_padrao()

    painel_antes = relatorio_service.obter_painel_inicio(
        usuario_admin_teste, data_inicio=hoje, data_fim=hoje
    )
    compra_service.registrar_compra(usuario_admin_teste, cliente.id, Decimal("77.00"), hoje, None)
    painel_depois = relatorio_service.obter_painel_inicio(
        usuario_admin_teste, data_inicio=hoje, data_fim=hoje
    )

    assert painel_depois.total_em_aberto_geral - painel_antes.total_em_aberto_geral == Decimal("77.00")
    assert painel_depois.periodo_inicio == hoje
    assert painel_depois.periodo_fim == hoje

    cliente_service.excluir_cliente(usuario_admin_teste, cliente.id)
