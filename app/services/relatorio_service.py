"""Serviço de relatórios, histórico e painel inicial.

Reúne consultas administrativas: histórico de alterações (auditoria), log
de erros do sistema, relatório de clientes com saldo em aberto (com
exportação para CSV) e o painel de início (maior valor gasto, mais contas
lançadas, evolução mensal e total em aberto geral).

Todas as funções exigem perfil Administrador, na mesma linha dos demais
serviços administrativos (usuários, backup, configurações).
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.database.connection import session_scope
from app.repositories import historico_repository, log_erro_repository, relatorio_repository
from app.services.auth_service import UsuarioAutenticado
from app.services.usuario_service import PermissaoNegadaError
from app.utils.error_handler import tratar_erros


@dataclass(frozen=True)
class HistoricoResumo:
    """Uma entrada do histórico de alterações, para exibição."""

    entidade: str
    entidade_id: str
    usuario_nome: str
    acao: str
    campo: Optional[str]
    valor_antigo: Optional[str]
    valor_novo: Optional[str]
    data_hora: datetime


@dataclass(frozen=True)
class LogErroResumo:
    """Uma entrada do log de erros, para exibição."""

    data_hora: datetime
    usuario: str
    erro: str
    stacktrace: Optional[str]


@dataclass(frozen=True)
class SaldoClienteResumo:
    """Saldo em aberto de um cliente, para o relatório."""

    id_visivel: int
    nome_principal: str
    total_em_aberto: Decimal


@dataclass(frozen=True)
class SaldoAtrasoResumo:
    """Saldo em aberto "atrasado" (compras em aberto há mais de N dias) de um cliente."""

    id: str
    id_visivel: int
    nome_principal: str
    telefone: Optional[str]
    total_em_atraso: Decimal
    dias_desde_a_compra_mais_antiga: int


def _exigir_administrador(usuario_logado: UsuarioAutenticado) -> None:
    if not usuario_logado.eh_administrador:
        raise PermissaoNegadaError("Apenas administradores podem acessar histórico e relatórios.")


@tratar_erros
def listar_historico(
    usuario_logado: UsuarioAutenticado, entidade: Optional[str] = None, limite: int = 200
) -> list[HistoricoResumo]:
    """Lista o histórico de alterações do sistema.

    Args:
        usuario_logado: Usuário autenticado que está consultando.
        entidade: Filtro opcional por entidade (ex.: "Cliente").
        limite: Número máximo de registros retornados.

    Returns:
        Lista de :class:`HistoricoResumo`, mais recente primeiro.

    Raises:
        PermissaoNegadaError: Se ``usuario_logado`` não for Administrador.
    """
    _exigir_administrador(usuario_logado)
    with session_scope() as session:
        registros = historico_repository.listar(session, entidade, limite)
        return [
            HistoricoResumo(
                entidade=r.entidade,
                entidade_id=str(r.entidade_id),
                usuario_nome=r.usuario.nome,
                acao=r.acao,
                campo=r.campo,
                valor_antigo=r.valor_antigo,
                valor_novo=r.valor_novo,
                data_hora=r.data_hora,
            )
            for r in registros
        ]


@tratar_erros
def listar_log_erros(usuario_logado: UsuarioAutenticado, limite: int = 200) -> list[LogErroResumo]:
    """Lista os erros registrados pelo sistema.

    Args:
        usuario_logado: Usuário autenticado que está consultando.
        limite: Número máximo de registros retornados.

    Returns:
        Lista de :class:`LogErroResumo`, mais recente primeiro.

    Raises:
        PermissaoNegadaError: Se ``usuario_logado`` não for Administrador.
    """
    _exigir_administrador(usuario_logado)
    with session_scope() as session:
        registros = log_erro_repository.listar(session, limite)
        return [
            LogErroResumo(
                data_hora=r.data_hora, usuario=r.usuario, erro=r.erro, stacktrace=r.stacktrace
            )
            for r in registros
        ]


@tratar_erros
def listar_saldos_em_aberto(usuario_logado: UsuarioAutenticado) -> list[SaldoClienteResumo]:
    """Lista os clientes com saldo em aberto (relatório).

    Args:
        usuario_logado: Usuário autenticado que está consultando.

    Returns:
        Lista de :class:`SaldoClienteResumo`, ordenada por nome principal.

    Raises:
        PermissaoNegadaError: Se ``usuario_logado`` não for Administrador.
    """
    _exigir_administrador(usuario_logado)
    with session_scope() as session:
        linhas = relatorio_repository.listar_saldos_em_aberto(session)
        return [
            SaldoClienteResumo(
                id_visivel=cliente.id_visivel,
                nome_principal=cliente.nome_principal,
                total_em_aberto=Decimal(total),
            )
            for cliente, total in linhas
        ]


@tratar_erros
def listar_saldos_em_atraso(
    usuario_logado: UsuarioAutenticado, dias_atraso: int = 30
) -> list[SaldoAtrasoResumo]:
    """Lista clientes com saldo "em atraso" (compras em aberto há mais de N dias).

    Usado para o envio de lembretes de cobrança. Como o fiado não tem uma
    data de vencimento própria, a data da compra é usada como referência —
    quanto mais antiga uma compra ainda em aberto, mais atrasada ela conta.

    Args:
        usuario_logado: Usuário autenticado que está consultando.
        dias_atraso: Quantidade de dias, a partir da data da compra, para
            considerá-la em atraso.

    Returns:
        Lista de :class:`SaldoAtrasoResumo`, da maior para a menor soma.

    Raises:
        PermissaoNegadaError: Se ``usuario_logado`` não for Administrador.
    """
    _exigir_administrador(usuario_logado)
    hoje = date.today()
    data_limite = hoje - timedelta(days=dias_atraso)

    with session_scope() as session:
        linhas = relatorio_repository.listar_saldos_em_atraso(session, data_limite)
        resultado = []
        for cliente, total, data_mais_antiga in linhas:
            telefone = cliente.telefones[0].numero if cliente.telefones else None
            resultado.append(
                SaldoAtrasoResumo(
                    id=str(cliente.id),
                    id_visivel=cliente.id_visivel,
                    nome_principal=cliente.nome_principal,
                    telefone=telefone,
                    total_em_atraso=Decimal(total),
                    dias_desde_a_compra_mais_antiga=(hoje - data_mais_antiga).days,
                )
            )
        return resultado


@tratar_erros
def exportar_saldos_em_aberto_csv(usuario_logado: UsuarioAutenticado, caminho_arquivo: str) -> None:
    """Exporta o relatório de saldo em aberto para um arquivo CSV.

    O arquivo usa ``;`` como separador e codificação UTF-8 com BOM, para
    abrir corretamente (acentos incluídos) no Excel.

    Args:
        usuario_logado: Usuário autenticado que está exportando.
        caminho_arquivo: Caminho completo do arquivo ``.csv`` a ser criado.

    Raises:
        PermissaoNegadaError: Se ``usuario_logado`` não for Administrador.
    """
    _exigir_administrador(usuario_logado)
    saldos = listar_saldos_em_aberto(usuario_logado)

    with open(caminho_arquivo, "w", newline="", encoding="utf-8-sig") as arquivo:
        escritor = csv.writer(arquivo, delimiter=";")
        escritor.writerow(["Código", "Cliente", "Total em Aberto (R$)"])
        for saldo in saldos:
            escritor.writerow(
                [saldo.id_visivel, saldo.nome_principal, f"{saldo.total_em_aberto:.2f}"]
            )


@tratar_erros
def exportar_saldos_em_aberto_xlsx(usuario_logado: UsuarioAutenticado, caminho_arquivo: str) -> None:
    """Exporta o relatório de saldo em aberto para uma planilha Excel (``.xlsx``) formatada.

    Diferente da exportação em CSV, a planilha já sai com cabeçalho em
    negrito, largura de coluna ajustada, valores em formato moeda e uma
    linha de total ao final.

    Args:
        usuario_logado: Usuário autenticado que está exportando.
        caminho_arquivo: Caminho completo do arquivo ``.xlsx`` a ser criado.

    Raises:
        PermissaoNegadaError: Se ``usuario_logado`` não for Administrador.
    """
    _exigir_administrador(usuario_logado)
    saldos = listar_saldos_em_aberto(usuario_logado)

    pasta_trabalho = Workbook()
    planilha = pasta_trabalho.active
    planilha.title = "Saldo em Aberto"

    cabecalhos = ["Código", "Cliente", "Total em Aberto (R$)"]
    planilha.append(cabecalhos)
    for coluna, _ in enumerate(cabecalhos, start=1):
        celula = planilha.cell(row=1, column=coluna)
        celula.font = Font(bold=True)
        celula.alignment = Alignment(horizontal="center")

    for saldo in saldos:
        planilha.append([saldo.id_visivel, saldo.nome_principal, float(saldo.total_em_aberto)])

    linha_total = planilha.max_row + 1
    planilha.cell(row=linha_total, column=1, value="Total")
    planilha.cell(row=linha_total, column=1).font = Font(bold=True)
    celula_total = planilha.cell(
        row=linha_total, column=3, value=float(sum((s.total_em_aberto for s in saldos), Decimal("0")))
    )
    celula_total.font = Font(bold=True)

    for linha in range(2, linha_total + 1):
        planilha.cell(row=linha, column=3).number_format = '"R$" #,##0.00'

    larguras = [10, 40, 20]
    for coluna, largura in enumerate(larguras, start=1):
        planilha.column_dimensions[get_column_letter(coluna)].width = largura

    planilha.freeze_panes = "A2"
    pasta_trabalho.save(caminho_arquivo)


@dataclass(frozen=True)
class ClienteValorResumo:
    """Um cliente e um valor associado (gasto ou saldo em aberto), para gráficos."""

    nome_principal: str
    valor: Decimal


@dataclass(frozen=True)
class ClienteContagemResumo:
    """Um cliente e uma quantidade associada (nº de contas lançadas), para gráficos."""

    nome_principal: str
    quantidade: int


@dataclass(frozen=True)
class PontoEvolucaoMensal:
    """Um ponto da série de evolução mensal de vendas."""

    mes: str  # "AAAA-MM"
    total: Decimal


@dataclass(frozen=True)
class PainelInicio:
    """Dados completos do painel de Início (dashboard), para um período."""

    periodo_inicio: date
    periodo_fim: date
    maior_valor_gasto: list[ClienteValorResumo]
    mais_contas_lancadas: list[ClienteContagemResumo]
    evolucao_mensal: list[PontoEvolucaoMensal]
    total_em_aberto_geral: Decimal
    maiores_saldos_em_aberto: list[ClienteValorResumo]


@tratar_erros
def obter_painel_inicio(
    usuario_logado: UsuarioAutenticado,
    data_inicio: Optional[date] = None,
    data_fim: Optional[date] = None,
) -> PainelInicio:
    """Monta os dados do painel de Início (dashboard) para o período informado.

    Args:
        usuario_logado: Usuário autenticado que está consultando.
        data_inicio: Início do período. Se omitido, usa o primeiro dia do
            mês atual.
        data_fim: Fim do período. Se omitido, usa a data de hoje.

    Returns:
        Um :class:`PainelInicio` com todos os indicadores.

    Raises:
        PermissaoNegadaError: Se ``usuario_logado`` não for Administrador.
    """
    if not usuario_logado.eh_administrador:
        raise PermissaoNegadaError("Apenas administradores podem ver o painel de início.")

    hoje = date.today()
    if data_inicio is None:
        data_inicio = hoje.replace(day=1)
    if data_fim is None:
        data_fim = hoje

    with session_scope() as session:
        maior_valor = relatorio_repository.listar_maior_valor_gasto(session, data_inicio, data_fim)
        mais_contas = relatorio_repository.listar_mais_contas_lancadas(session, data_inicio, data_fim)
        evolucao = relatorio_repository.listar_evolucao_mensal(session, meses=6)
        total_aberto_geral = relatorio_repository.calcular_total_em_aberto_geral(session)
        saldos = relatorio_repository.listar_saldos_em_aberto(session)

        maiores_saldos = sorted(
            (
                ClienteValorResumo(nome_principal=cliente.nome_principal, valor=Decimal(total))
                for cliente, total in saldos
            ),
            key=lambda item: item.valor,
            reverse=True,
        )[:10]

        return PainelInicio(
            periodo_inicio=data_inicio,
            periodo_fim=data_fim,
            maior_valor_gasto=[
                ClienteValorResumo(nome_principal=cliente.nome_principal, valor=Decimal(valor))
                for cliente, valor in maior_valor
            ],
            mais_contas_lancadas=[
                ClienteContagemResumo(nome_principal=cliente.nome_principal, quantidade=int(quantidade))
                for cliente, quantidade in mais_contas
            ],
            evolucao_mensal=[PontoEvolucaoMensal(mes=mes, total=total) for mes, total in evolucao],
            total_em_aberto_geral=total_aberto_geral,
            maiores_saldos_em_aberto=maiores_saldos,
        )
